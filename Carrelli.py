import streamlit as st
from pathlib import Path
import openpyxl
import base64
import html
import re


# ==========================================================
# CONFIGURAZIONE
# ==========================================================

BASE_DIR = Path(__file__).resolve().parent

FILE_EXCEL = BASE_DIR / "ATTIVITA' CARRELLO.xlsm"


# ==========================================================
# CSS
# ==========================================================

st.markdown("""
<style>

.carrelli-header {
    background: linear-gradient(90deg, #d40000, #ed1c24);
    color: white;
    padding: 16px 22px;
    border-radius: 10px;
    text-align: center;
    font-size: 26px;
    font-weight: bold;
    margin-bottom: 20px;
}


/* ======================================================
   CONTENITORE FOGLIO
   ====================================================== */

.excel-wrapper {
    width: 100%;
    overflow: auto;
    background: white;
    border-radius: 8px;
    padding: 10px;
}


/* ======================================================
   TABELLA
   ====================================================== */

.excel-table {
    border-collapse: collapse;
    table-layout: fixed;
    background: white;
}


/*
   IMPORTANTISSIMO:

   NESSUN BORDO GENERALE.

   Questo elimina il reticolato.
*/

.excel-table td {
    border: none !important;
    padding: 0;
    margin: 0;
    vertical-align: middle;
    overflow: hidden;
    background-clip: padding-box;
}


/* ======================================================
   CONTENUTO CELLA
   ====================================================== */

.excel-cell-content {
    width: 100%;
    height: 100%;
    box-sizing: border-box;
    overflow: hidden;
}


/* ======================================================
   TESTO
   ====================================================== */

.excel-text {
    width: 100%;
    box-sizing: border-box;
}


/* ======================================================
   IMMAGINI
   ====================================================== */

.excel-table img {
    display: block;
    max-width: 100%;
    height: auto;
    margin-left: auto;
    margin-right: auto;
}


/* ======================================================
   IMMAGINI GRANDI
   ====================================================== */

.excel-image {
    display: block;
    max-width: 100%;
    max-height: 500px;
    object-fit: contain;
}


/* ======================================================
   LINK
   ====================================================== */

.excel-link {
    color: #0066cc;
    font-weight: bold;
    text-decoration: none;
}


/* ======================================================
   SCROLL ORIZZONTALE
   ====================================================== */

.excel-wrapper::-webkit-scrollbar {
    height: 10px;
    width: 10px;
}

.excel-wrapper::-webkit-scrollbar-thumb {
    background: #aaa;
    border-radius: 5px;
}


/* ======================================================
   SELETTORI
   ====================================================== */

.carrelli-select label {
    font-weight: bold;
}

</style>
""", unsafe_allow_html=True)


# ==========================================================
# SEZIONI
# ==========================================================

SEZIONI = {

    "🛞 CARRELLI": [
        "DM1-CARR.1",
        "DM1-CARR.2",
        "M3-CARR.1",
        "M3-CARR.2",
        "M6-CARR.1",
        "M6-CARR.2",
        "DM8-CARR.1",
        "DM8-CARR.2",
    ],

    "📡 SENSORI": [
        "SENSORI SPM",
        "PT100 RIDUTTORI",
    ],

    "🔌 FUSE LOOP": [
        "FUSE LOOP CASSA MOTOR",
        "FUSE LOOP TRENO COMPLETO",
    ],

    "🔄 DNRA": [
        "LOOP DNRA",
        "OVERVIEW DNRA",
    ],

    "🚆 STATO TRENO": [
        "STATO TRENO",
    ],
}


# ==========================================================
# FUNZIONI COLORE
# ==========================================================

def colore_excel(colore):

    if colore is None:
        return None

    try:

        rgb = colore.rgb

        if rgb:

            rgb = str(rgb)

            # AARRGGBB → RRGGBB
            if len(rgb) == 8:
                rgb = rgb[2:]

            if len(rgb) == 6:
                return "#" + rgb

    except Exception:
        pass

    return None


# ==========================================================
# FONT
# ==========================================================

def font_css(font):

    css = []

    if not font:
        return ""

    try:

        if font.bold:
            css.append("font-weight:bold;")

        if font.italic:
            css.append("font-style:italic;")

        if font.sz:
            css.append(
                f"font-size:{float(font.sz):.1f}pt;"
            )

        if font.name:
            css.append(
                f"font-family:'{html.escape(str(font.name))}';"
            )

        colore = colore_excel(font.color)

        if colore:
            css.append(
                f"color:{colore};"
            )

        if font.underline:
            css.append(
                "text-decoration:underline;"
            )

    except Exception:
        pass

    return "".join(css)


# ==========================================================
# SFONDO
# ==========================================================

def fill_css(fill):

    if not fill:
        return ""

    try:

        if fill.fill_type:

            colore = colore_excel(
                fill.fgColor
            )

            if colore:
                return (
                    f"background-color:{colore};"
                )

    except Exception:
        pass

    return ""


# ==========================================================
# ALLINEAMENTO
# ==========================================================

def alignment_css(alignment):

    css = []

    if not alignment:
        return ""

    try:

        if alignment.horizontal:

            valore = str(
                alignment.horizontal
            )

            if valore == "centerContinuous":
                valore = "center"

            if valore in [
                "left",
                "center",
                "right",
                "justify"
            ]:

                css.append(
                    f"text-align:{valore};"
                )

        if alignment.vertical:

            valore = str(
                alignment.vertical
            )

            if valore in [
                "top",
                "center",
                "bottom"
            ]:

                if valore == "center":
                    valore = "middle"

                css.append(
                    f"vertical-align:{valore};"
                )

        if alignment.wrap_text:

            css.append(
                "white-space:normal;"
            )

        else:

            css.append(
                "white-space:nowrap;"
            )

    except Exception:
        pass

    return "".join(css)


# ==========================================================
# BORDI
# ==========================================================

def bordo_reale_css(border, mostra=False):

    """
    Di default NON mostra i bordi Excel.

    Questo è ciò che elimina il reticolato.

    Se mostra=True, visualizza solo i bordi
    realmente presenti nella cella.
    """

    if not mostra:
        return ""

    css = []

    lati = [
        ("top", border.top),
        ("bottom", border.bottom),
        ("left", border.left),
        ("right", border.right)
    ]

    for nome, lato in lati:

        try:

            if lato and lato.style:

                colore = colore_excel(
                    lato.color
                )

                if not colore:
                    colore = "#777777"

                css.append(
                    f"border-{nome}:1px solid {colore};"
                )

        except Exception:
            pass

    return "".join(css)


# ==========================================================
# LARGHEZZA COLONNA
# ==========================================================

def larghezza_colonna(ws, col):

    lettera = openpyxl.utils.get_column_letter(
        col
    )

    dimensione = ws.column_dimensions[
        lettera
    ].width

    if dimensione is None:
        dimensione = 8.43

    try:

        pixel = int(
            float(dimensione) * 7
        )

    except Exception:

        pixel = 60

    return max(
        pixel,
        25
    )


# ==========================================================
# ALTEZZA RIGA
# ==========================================================

def altezza_riga(ws, row):

    altezza = ws.row_dimensions[
        row
    ].height

    if altezza is None:
        altezza = 15

    try:

        pixel = int(
            float(altezza) * 1.33
        )

    except Exception:

        pixel = 20

    return max(
        pixel,
        18
    )


# ==========================================================
# IMMAGINI EXCEL
# ==========================================================

def estrai_immagini(ws):

    immagini = {}

    try:

        for img in ws._images:

            try:

                anchor = img.anchor

                if not hasattr(
                    anchor,
                    "_from"
                ):
                    continue

                col = (
                    anchor._from.col
                    + 1
                )

                row = (
                    anchor._from.row
                    + 1
                )

                # --------------------------------------------------
                # DATI IMMAGINE
                # --------------------------------------------------

                image_bytes = img._data()

                if not image_bytes:
                    continue

                # --------------------------------------------------
                # FORMATO
                # --------------------------------------------------

                formato = "png"

                try:

                    if hasattr(
                        img,
                        "format"
                    ):

                        if img.format:

                            formato = str(
                                img.format
                            ).lower()

                except Exception:
                    pass

                if formato == "jpg":
                    formato = "jpeg"

                # --------------------------------------------------
                # BASE64
                # --------------------------------------------------

                encoded = base64.b64encode(
                    image_bytes
                ).decode(
                    "utf-8"
                )

                # --------------------------------------------------
                # DIMENSIONI
                # --------------------------------------------------

                larghezza = None
                altezza = None

                try:

                    if hasattr(
                        anchor,
                        "ext"
                    ):

                        # EMU → pixel circa
                        larghezza = int(
                            anchor.ext.cx
                            / 9525
                        )

                        altezza = int(
                            anchor.ext.cy
                            / 9525
                        )

                except Exception:
                    pass

                immagini[
                    (row, col)
                ] = {

                    "data": encoded,

                    "format": formato,

                    "width": larghezza,

                    "height": altezza

                }

            except Exception:
                continue

    except Exception as e:

        st.warning(
            "⚠️ Alcune immagini del foglio "
            f"non sono state lette: {e}"
        )

    return immagini


# ==========================================================
# MAPPA CELLE UNITE
# ==========================================================

def crea_mappa_merge(ws):

    merged_map = {}

    for merged in ws.merged_cells.ranges:

        min_col = merged.min_col
        max_col = merged.max_col

        min_row = merged.min_row
        max_row = merged.max_row

        merged_map[
            (min_row, min_col)
        ] = (

            max_row - min_row + 1,

            max_col - min_col + 1

        )

        for row in range(
            min_row,
            max_row + 1
        ):

            for col in range(
                min_col,
                max_col + 1
            ):

                if (
                    row != min_row
                    or col != min_col
                ):

                    merged_map[
                        (row, col)
                    ] = "skip"

    return merged_map


# ==========================================================
# TESTO
# ==========================================================

def crea_contenuto_testo(valore):

    if valore is None:
        return ""

    testo = str(valore)

    if not testo.strip():
        return ""

    # ------------------------------------------------------
    # LINK
    # ------------------------------------------------------

    if testo.startswith(
        "http://"
    ) or testo.startswith(
        "https://"
    ):

        url = html.escape(
            testo,
            quote=True
        )

        return (
            f'<a class="excel-link" '
            f'href="{url}" '
            f'target="_blank">'
            f'📄 {html.escape(testo)}'
            f'</a>'
        )

    # ------------------------------------------------------
    # TESTO NORMALE
    # ------------------------------------------------------

    testo = html.escape(
        testo
    )

    # Mantiene gli a capo
    testo = testo.replace(
        "\n",
        "<br>"
    )

    return (
        f'<div class="excel-text">'
        f'{testo}'
        f'</div>'
    )


# ==========================================================
# RENDER FOGLIO
# ==========================================================

def render_foglio(ws):

    immagini = estrai_immagini(
        ws
    )

    merged_map = crea_mappa_merge(
        ws
    )

    # ======================================================
    # DIMENSIONI
    # ======================================================

    max_row = ws.max_row
    max_col = ws.max_column

    # ======================================================
    # HTML
    # ======================================================

    html_out = []

    html_out.append(
        '<div class="excel-wrapper">'
    )

    html_out.append(
        '<table class="excel-table">'
    )

    # ======================================================
    # COLONNE
    # ======================================================

    html_out.append(
        "<colgroup>"
    )

    for col in range(
        1,
        max_col + 1
    ):

        width = larghezza_colonna(
            ws,
            col
        )

        html_out.append(
            f'<col style="width:{width}px;">'
        )

    html_out.append(
        "</colgroup>"
    )

    # ======================================================
    # RIGHE
    # ======================================================

    for row in range(
        1,
        max_row + 1
    ):

        altezza = altezza_riga(
            ws,
            row
        )

        html_out.append(
            f'<tr style="height:{altezza}px;">'
        )

        for col in range(
            1,
            max_col + 1
        ):

            # ------------------------------------------------
            # CELLA MERGED SECONDARIA
            # ------------------------------------------------

            if merged_map.get(
                (row, col)
            ) == "skip":

                continue

            cella = ws.cell(
                row=row,
                column=col
            )

            valore = cella.value

            img_info = immagini.get(
                (row, col)
            )

            # ------------------------------------------------
            # CONTENUTO
            # ------------------------------------------------

            contenuto = ""

            # TESTO

            if valore is not None:

                contenuto += (
                    crea_contenuto_testo(
                        valore
                    )
                )

            # ------------------------------------------------
            # IMMAGINE
            # ------------------------------------------------

            if img_info:

                formato = img_info[
                    "format"
                ]

                data = img_info[
                    "data"
                ]

                width = img_info.get(
                    "width"
                )

                height = img_info.get(
                    "height"
                )

                stile_img = (
                    "max-width:100%;"
                    "height:auto;"
                )

                if width:
                    stile_img += (
                        f"width:{width}px;"
                    )

                if height:
                    stile_img += (
                        f"max-height:{height}px;"
                    )

                contenuto += (
                    f'<img '
                    f'class="excel-image" '
                    f'src="data:image/{formato};'
                    f'base64,{data}" '
                    f'style="{stile_img}">'
                )

            # ------------------------------------------------
            # STILE
            # ------------------------------------------------

            stile = ""

            stile += fill_css(
                cella.fill
            )

            stile += font_css(
                cella.font
            )

            stile += alignment_css(
                cella.alignment
            )

            # ------------------------------------------------
            # BORDI
            #
            # IMPORTANTE:
            #
            # FALSE = niente reticolato
            # ------------------------------------------------

            stile += bordo_reale_css(
                cella.border,
                mostra=False
            )

            # ------------------------------------------------
            # MERGE
            # ------------------------------------------------

            rowspan = 1
            colspan = 1

            merge_info = merged_map.get(
                (row, col)
            )

            if isinstance(
                merge_info,
                tuple
            ):

                rowspan = merge_info[0]
                colspan = merge_info[1]

            # ------------------------------------------------
            # TD
            # ------------------------------------------------

            html_out.append(
                "<td "
                f'style="{stile}"'
            )

            if rowspan > 1:

                html_out.append(
                    f' rowspan="{rowspan}"'
                )

            if colspan > 1:

                html_out.append(
                    f' colspan="{colspan}"'
                )

            html_out.append(">")

            html_out.append(
                '<div class="excel-cell-content">'
            )

            html_out.append(
                contenuto
            )

            html_out.append(
                "</div>"
            )

            html_out.append(
                "</td>"
            )

        html_out.append(
            "</tr>"
        )

    html_out.append(
        "</table>"
    )

    html_out.append(
        "</div>"
    )

    return "".join(
        html_out
    )


# ==========================================================
# PAGINA CARRELLI
# ==========================================================

def carrelli_page():

    # ======================================================
    # TITOLO
    # ======================================================

    st.markdown(
        """
        <div class="carrelli-header">
            🚆 CARRELLI ETR1000
        </div>
        """,
        unsafe_allow_html=True
    )

    # ======================================================
    # FILE
    # ======================================================

    if not FILE_EXCEL.exists():

        st.error(
            "❌ File Excel non trovato."
        )

        st.code(
            str(FILE_EXCEL)
        )

        st.info(
            """
            Il file deve essere presente
            nella stessa cartella di Carrelli.py:

            ATTIVITA' CARRELLO.xlsm
            """
        )

        return

    # ======================================================
    # APERTURA EXCEL
    # ======================================================

    try:

        wb = openpyxl.load_workbook(
            FILE_EXCEL,
            read_only=False,
            data_only=False,
            keep_vba=True
        )

    except Exception as e:

        st.error(
            "❌ Errore apertura Excel."
        )

        st.code(
            str(e)
        )

        return

    # ======================================================
    # FOGLI
    # ======================================================

    fogli_disponibili = wb.sheetnames

    # ======================================================
    # SEZIONI DISPONIBILI
    # ======================================================

    sezioni_disponibili = {}

    for nome_sezione, fogli in SEZIONI.items():

        presenti = [

            f

            for f in fogli

            if f in fogli_disponibili

        ]

        if presenti:

            sezioni_disponibili[
                nome_sezione
            ] = presenti

    if not sezioni_disponibili:

        st.error(
            "❌ Nessuno dei fogli configurati "
            "è presente nel file Excel."
        )

        st.write(
            "Fogli trovati nel file:"
        )

        st.write(
            fogli_disponibili
        )

        wb.close()

        return

    # ======================================================
    # SEZIONE
    # ======================================================

    st.markdown(
        "### 📂 Sezione"
    )

    sezione = st.selectbox(
        "Sezione",
        list(
            sezioni_disponibili.keys()
        ),
        label_visibility="collapsed"
    )

    # ======================================================
    # FOGLIO
    # ======================================================

    st.markdown(
        "### 📄 Foglio"
    )

    foglio = st.selectbox(
        "Foglio",
        sezioni_disponibili[
            sezione
        ],
        label_visibility="collapsed"
    )

    st.divider()

    # ======================================================
    # FOGLIO EXCEL
    # ======================================================

    ws = wb[
        foglio
    ]

    # ======================================================
    # NASCONDI GRIDLINES
    # ======================================================

    try:

        ws.sheet_view.showGridLines = False

    except Exception:
        pass

    # ======================================================
    # INFORMAZIONI
    # ======================================================

    st.markdown(
        f"""
        <div style="
            background:#f3f3f3;
            padding:10px 14px;
            border-radius:8px;
            margin-bottom:15px;
        ">
            📂 <b>{html.escape(sezione)}</b>
            &nbsp;&nbsp; | &nbsp;&nbsp;
            📄 <b>{html.escape(foglio)}</b>
        </div>
        """,
        unsafe_allow_html=True
    )

    # ======================================================
    # RENDER
    # ======================================================

    try:

        with st.spinner(
            "🔄 Caricamento schema..."
        ):

            contenuto = render_foglio(
                ws
            )

        st.markdown(
            contenuto,
            unsafe_allow_html=True
        )

    except Exception as e:

        st.error(
            "❌ Errore nella visualizzazione."
        )

        st.code(
            str(e)
        )

    finally:

        try:
            wb.close()
        except Exception:
            pass


# ==========================================================
# AVVIO DIRETTO
# ==========================================================

if __name__ == "__main__":

    carrelli_page()
